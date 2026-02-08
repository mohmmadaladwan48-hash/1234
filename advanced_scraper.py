import instaloader
import time
import os
import json
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ensure UTF-8 encoding for proper Arabic text display
import sys
import io
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class InstagramInfoScraper:
    """Advanced Instagram scraper with session management and export features"""
    
    def __init__(self):
        self.session_dir = Path("./sessions")
        self.session_dir.mkdir(exist_ok=True)
        self.output_dir = Path("./output")
        self.output_dir.mkdir(exist_ok=True)
        self.search_history_file = self.output_dir / "search_history.json"
        self.loader = None
        self.current_user = None
        self.search_history = self.load_search_history()
        
    def load_search_history(self):
        """Load previous search history from file"""
        if self.search_history_file.exists():
            try:
                with open(self.search_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}
        
    def create_loader(self):
        """Create Instaloader instance with optimized settings"""
        return instaloader.Instaloader(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            quiet=False,
            sleep=True
        )
    
    def login(self, username, password, force_relogin=False):
        """Login to Instagram with session management"""
        self.loader = self.create_loader()
        session_file = self.session_dir / username
        
        try:
            if session_file.exists() and not force_relogin:
                print(f"🔄 Loading saved session for {username}...")
                try:
                    self.loader.load_session_from_file(username, filename=str(session_file))
                    print("✅ Session restored!")
                    self.current_user = username
                    return True
                except Exception as e:
                    print(f"⚠️ Session restore failed, logging in again...")
                    self.loader.login(username, password)
                    print("✅ Login successful!")
                    self.loader.save_session_to_file(filename=str(session_file))
                    self.current_user = username
                    return True
            else:
                print(f"🔐 Logging in as {username}...")
                self.loader.login(username, password)
                print("✅ Login successful!")
                self.loader.save_session_to_file(filename=str(session_file))
                self.current_user = username
                return True
                
        except Exception as e:
            print(f"❌ Login failed: {e}")
            return False
    
    def get_user_info(self, username, retries=5, delay=10):
        """Fetch detailed user information"""
        if not self.loader:
            self.loader = self.create_loader()
        
        attempt = 0
        while attempt < retries:
            try:
                profile = instaloader.Profile.from_username(self.loader.context, username)
                
                # Extract location data (if available)
                country = "Not available"
                city = "Not available"
                
                try:
                    if hasattr(profile, 'business_address_json') and profile.business_address_json:
                        country = profile.business_address_json.get("country") or "Not available"
                        city = profile.business_address_json.get("city") or "Not available"
                except (AttributeError, TypeError):
                    pass
                
                full_location = f"{city}, {country}" if city != "Not available" and country != "Not available" else city if city != "Not available" else country
                
                # Count name changes
                name_changes = "Not available"
                if hasattr(profile, 'name_changes') and profile.name_changes:
                    name_changes = len(profile.name_changes)
                
                data = {
                    "username": profile.username,
                    "full_name": profile.full_name,
                    "followers": f"{profile.followers:,}",
                    "following": f"{profile.followees:,}",
                    "bio": profile.biography if profile.biography else "No bio set",
                    "city": city,
                    "country": country,
                    "full_location": full_location,
                    "posts_count": profile.mediacount,
                    "name_changes": name_changes,
                    "is_business_account": "✅ Yes" if profile.is_business_account else "❌ No",
                    "is_verified": "✅ Yes" if profile.is_verified else "❌ No",
                    "is_public": "🌐 Yes" if not profile.is_private else "🔒 No",
                    "external_url": profile.external_url if profile.external_url else "Not set",
                    "profile_pic_url": profile.profile_pic_url,
                    "biography_html": profile.biography_html if hasattr(profile, 'biography_html') else "Not available",
                    "search_timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Save to search history
                self.search_history[username.lower()] = data
                self.save_search_history()
                
                return data
                
            except Exception as e:
                error_msg = str(e)
                if any(x in error_msg for x in ["401", "Unauthorized", "Please wait", "429", "Too Many Requests"]):
                    attempt += 1
                    if attempt < retries:
                        wait_time = delay * attempt
                        print(f"⏳ Rate limited. Waiting {wait_time} seconds before retry {attempt}/{retries-1}...")
                        time.sleep(wait_time)
                    else:
                        return f"❌ Rate limited by Instagram. Please wait 30-60 minutes or login with your account."
                elif "does not exist" in error_msg.lower() or "not found" in error_msg.lower():
                    return f"❌ User '@{username}' not found on Instagram."
                else:
                    return f"❌ Error: {e}"
        
        return f"❌ Failed after {retries} attempts."
    
    def export_to_json(self, data, username):
        """Export user data to JSON file"""
        filename = self.output_dir / f"{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return filename
    
    def export_to_csv(self, data_list, filename="instagram_accounts.csv"):
        """Export multiple user data to CSV file"""
        filepath = self.output_dir / filename
        
        if not data_list:
            return None
        
        # Filter to get only dictionary results
        valid_data = [d for d in data_list if isinstance(d, dict)]
        
        if not valid_data:
            return None
        
        keys = valid_data[0].keys()
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(valid_data)
        
        return filepath
    
    def save_search_history(self):
        """Save search history to JSON file"""
        with open(self.search_history_file, 'w', encoding='utf-8') as f:
            json.dump(self.search_history, f, indent=2, ensure_ascii=False)
    
    def get_all_searches(self):
        """Get all previous searches"""
        return self.search_history
    
    def export_search_history_to_csv(self, filename="all_searches.csv"):
        """Export all search history to CSV"""
        if not self.search_history:
            return None
        
        filepath = self.output_dir / filename
        data_list = list(self.search_history.values())
        
        keys = data_list[0].keys()
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(data_list)
        
        return filepath
    
    def export_search_history_to_excel(self, filename="instagram_search_results.xlsx"):
        """Export all search history to Excel file with professional formatting"""
        if not self.search_history:
            return None
        
        filepath = self.output_dir / filename
        
        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram Users"
        
        # Define professional styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=13, name='Calibri')
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Alternating row colors
        row_fill_light = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Data styling
        data_font = Font(size=11, name='Calibri')
        data_alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        data_alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Get all unique keys
        data_list = list(self.search_history.values())
        if not data_list:
            return None
        
        keys = list(data_list[0].keys())
        
        # Define professional column widths
        column_widths = {
            'username': 18,
            'full_name': 28,
            'followers': 18,
            'following': 18,
            'bio': 40,
            'city': 18,
            'country': 18,
            'full_location': 28,
            'posts_count': 14,
            'name_changes': 14,
            'is_business_account': 18,
            'is_verified': 18,
            'is_public': 18,
            'external_url': 45,
            'profile_pic_url': 45,
            'biography_html': 45,
            'search_timestamp': 20
        }
        
        # Write headers with professional styling
        for col_num, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = key.replace("_", " ").title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_thin
            
            # Set column width
            width = column_widths.get(key, 20)
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Write data with alternating colors
        for row_num, data in enumerate(data_list, 2):
            # Alternate row colors
            is_even_row = (row_num - 2) % 2 == 0
            row_fill = row_fill_white if is_even_row else row_fill_light
            
            for col_num, key in enumerate(keys, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = data.get(key, "N/A")
                cell.value = value
                cell.font = data_font
                cell.border = border_thin
                cell.fill = row_fill
                
                # Set alignment based on content type
                if key in ['full_name', 'bio', 'biography_html']:
                    # Arabic text - right aligned
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                elif key in ['followers', 'following', 'posts_count', 'name_changes']:
                    # Numbers - center aligned
                    cell.alignment = data_alignment_center
                elif key.startswith('is_'):
                    # Boolean values - center aligned
                    cell.alignment = data_alignment_center
                else:
                    # Text - left aligned
                    cell.alignment = data_alignment_left
        
        # Set row height for header and data
        ws.row_dimensions[1].height = 30
        for row in range(2, len(data_list) + 2):
            ws.row_dimensions[row].height = 25
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add autofilter to headers
        ws.auto_filter.ref = f"A1:{chr(64 + len(keys))}{len(data_list) + 1}"
        
        # Save workbook
        wb.save(filepath)
        
        return filepath


def display_info(info):
    """Display user information in formatted table"""
    if isinstance(info, dict):
        print("\n✅ SUCCESS! Here's the account information:\n")
        for key, value in info.items():
            key_display = key.replace("_", " ").title()
            print(f"  {key_display:<25}: {value}")
        return True
    else:
        print(f"\n{info}")
        return False


def main():
    print("=" * 70)
    print("📸 INSTAGRAM ACCOUNT INFO SCRAPER (Advanced)")
    print("=" * 70)
    print("\n✨ Features: No login required | Batch lookup | Arabic support | Auto-save all searches")
    print("💾 All searches are automatically saved to /output folder\n")
    
    scraper = InstagramInfoScraper()
    
    # Optional login prompt
    login_choice = input("Do you want to login for better results? (yes/no): ").strip().lower()
    
    if login_choice in ['yes', 'y']:
        username = input("Enter your Instagram username: ").strip()
        password = input("Enter your Instagram password: ").strip()
        
        if not scraper.login(username, password):
            print("⚠️ Continuing without login (will still work for public accounts)...\n")
    else:
        print("⚠️ Continuing without login (public accounts only)...\n")
    
    # Main menu loop
    while True:
        print("\n" + "=" * 70)
        print("📋 MAIN MENU")
        print("=" * 70)
        print("  1️⃣  Lookup single account")
        print("  2️⃣  Batch lookup (multiple accounts)")
        print("  3️⃣  View search history")
        print("  4️⃣  Export all searches to CSV")
        print("  5️⃣  Change login account")
        print("  6️⃣  Exit / Quit")
        print("=" * 70)
        
        choice = input("\n👉 Choose option (1-6) or press 'q' to quit: ").strip().lower()
        
        # Handle quit options
        if choice in ['q', 'quit', '6']:
            # Auto-generate Excel file on quit
            if scraper.search_history:
                print("\n" + "=" * 70)
                print("⏳ Generating Excel file with all searches...")
                excel_path = scraper.export_search_history_to_excel()
                if excel_path:
                    print(f"✅ Excel file created: {excel_path}")
            
            print("\n" + "=" * 70)
            print(f"📊 Total searches saved: {len(scraper.search_history)}")
            print(f"📁 All data saved to: {scraper.output_dir}")
            print("👋 Thank you for using Instagram Scraper! Goodbye!\n")
            break
        
        if choice == "1":
            print("\n" + "-" * 70)
            username = input("📱 Enter Instagram username: ").strip()
            if not username:
                print("❌ Username cannot be empty!")
                continue
            
            print("⏳ Fetching information...")
            info = scraper.get_user_info(username)
            
            if display_info(info):
                # Export option
                export_choice = input("\n💾 Export this user to JSON? (yes/no): ").strip().lower()
                if export_choice in ['yes', 'y']:
                    filepath = scraper.export_to_json(info, username)
                    print(f"✅ Exported to: {filepath}")
        
        elif choice == "2":
            print("\n" + "-" * 70)
            usernames_input = input("📱 Enter usernames (comma-separated, e.g., user1,user2,user3): ").strip()
            if not usernames_input:
                print("❌ No usernames provided!")
                continue
            
            usernames = [u.strip() for u in usernames_input.split(',')]
            
            print(f"\n🔍 Starting batch lookup for {len(usernames)} accounts...\n")
            
            results = []
            for i, user in enumerate(usernames, 1):
                print(f"[{i}/{len(usernames)}] 📊 Fetching @{user}...", end=" ", flush=True)
                info = scraper.get_user_info(user)
                if isinstance(info, dict):
                    results.append(info)
                    print("✅")
                else:
                    print("❌")
                    print(f"      {info}\n")
            
            # Export batch results
            if results:
                print(f"\n✅ Successfully fetched {len(results)}/{len(usernames)} accounts")
                export_choice = input(f"💾 Export {len(results)} results to CSV? (yes/no): ").strip().lower()
                if export_choice in ['yes', 'y']:
                    filepath = scraper.export_to_csv(results)
                    print(f"✅ Exported to: {filepath}")
            else:
                print("\n❌ No accounts were successfully fetched.")
        
        elif choice == "3":
            print("\n" + "-" * 70)
            if scraper.search_history:
                print(f"📋 Search History ({len(scraper.search_history)} users):\n")
                for i, (username, data) in enumerate(scraper.search_history.items(), 1):
                    full_name = data.get('full_name', 'N/A')
                    followers = data.get('followers', 'N/A')
                    timestamp = data.get('search_timestamp', 'N/A')
                    print(f"  {i}. @{username} ({full_name}) - {followers} followers - {timestamp}")
            else:
                print("❌ No search history found.")
        
        elif choice == "4":
            print("\n" + "-" * 70)
            if scraper.search_history:
                print("💾 Export options:")
                print("  1. Export to CSV")
                print("  2. Export to Excel")
                print("  3. Export both")
                export_type = input("Choose export type (1-3): ").strip()
                
                if export_type in ['1', '3']:
                    csv_path = scraper.export_search_history_to_csv()
                    if csv_path:
                        print(f"✅ CSV file created: {csv_path}")
                
                if export_type in ['2', '3']:
                    excel_path = scraper.export_search_history_to_excel()
                    if excel_path:
                        print(f"✅ Excel file created: {excel_path}")
            else:
                print("❌ No searches to export.")
        
        elif choice == "5":
            print("\n" + "-" * 70)
            username = input("📱 Enter your Instagram username: ").strip()
            password = input("🔐 Enter your Instagram password: ").strip()
            
            if scraper.login(username, password, force_relogin=True):
                print("✅ Successfully logged in with new account!")
            else:
                print("❌ Login failed. Check your credentials and try again.")
        
        else:
            print("\n❌ Invalid option. Please choose 1-6 or press 'q' to quit.")


if __name__ == "__main__":
    main()


if __name__ == "__main__":
    main()
